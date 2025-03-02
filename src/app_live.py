import streamlit as st
import cv2
import torch
import openpyxl
import pandas as pd
from datetime import datetime
from pathlib import Path
from ultralytics import YOLO
from helper import is_face_present_for_numpy, who, crop_face_from_numpy, reader

# Streamlit Page Setup
st.set_page_config(page_title="Live Face Tracking", layout="wide")
st.sidebar.title("Navigation")
page = st.sidebar.radio("Go to", ["Live Camera", "Attendance Log", "Settings"])

# Set up device (MPS for Mac, CUDA for GPU, else CPU)
if torch.backends.mps.is_available():
    device = torch.device("mps")
elif torch.cuda.is_available():
    device = torch.device("cuda")
else:
    device = torch.device("cpu")

# Load YOLO model
yolo_model_path = Path("models/yolo/yolov8_77_epochs.pt")
yolo_model = YOLO(yolo_model_path)

# Attendance file path
attendance_file = Path("outputs/attendance_results/tracked_objects.xlsx")
if attendance_file.exists():
    workbook = openpyxl.load_workbook(attendance_file)
    sheet = workbook.active
else:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Tracking Log"
    sheet.append(["Person Name", "Date", "Time", "Last Time", "Location"])
    workbook.save(attendance_file)

# Store logged names to avoid duplicates
logged_names = set()
for row in sheet.iter_rows(min_row=2, values_only=True):  
    if row[0] is not None:
        logged_names.add(row[0])

# Page 1: Live Camera
if page == "Live Camera":
    st.header("Live Face Tracking")
    frame_placeholder = st.empty()

    cap = cv2.VideoCapture(0)  # Open webcam

    if not cap.isOpened():
        st.error("Could not access webcam. Please check your camera.")
        st.stop()

    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            st.error("Failed to capture frame from webcam.")
            break

        original_frame = frame.copy()
        results = yolo_model.track(frame, conf=0.5, iou=0.1, persist=True, device=device, imgsz=1024)

        detected_person = "Unknown"
        
        for result in results:
            boxes = result.boxes.xyxy if result.boxes.xyxy is not None else []
            ids = result.boxes.id if result.boxes.id is not None else []
            classes = result.boxes.cls if result.boxes.cls is not None else []

            for box, obj_id, cls in zip(boxes, ids, classes):
                if obj_id is None:
                    continue

                x1, y1, x2, y2 = map(int, box)
                class_name = yolo_model.names[int(cls)]

                cropped_image = original_frame[y1:y2, x1:x2]

                if is_face_present_for_numpy(cropped_image):
                    cropped_image = crop_face_from_numpy(cropped_image)
                    detected_person = who(cropped_image)

                    # Read Date, Time, and Location
                    result = reader.readtext(original_frame)
                    date, time_value, location = None, None, None
                    if len(result) >= 4:
                        date, time_value, location = result[0][1], result[2][1], result[3][1]

                    # Log Attendance
                    if detected_person not in logged_names:
                        logged_names.add(detected_person)
                        sheet.append([detected_person, date, time_value, time_value, location])
                        workbook.save(attendance_file)

                    # Draw Bounding Box & ID Label
                    cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                    cv2.putText(
                        frame,
                        f"{class_name} ID:{detected_person}",
                        (x1, y1 - 10),
                        cv2.FONT_HERSHEY_SIMPLEX,
                        0.5,
                        (0, 255, 0),
                        2,
                    )

        # Convert BGR to RGB and update Streamlit display
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        frame_placeholder.image(frame_rgb, channels="RGB")

    cap.release()
    cv2.destroyAllWindows()

# Page 2: Attendance Log
elif page == "Attendance Log":
    st.header("Attendance Log")
    if attendance_file.exists():
        df = pd.read_excel(attendance_file)
        st.dataframe(df)
    else:
        st.write("No attendance records found.")


