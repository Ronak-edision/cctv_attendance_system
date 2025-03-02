import streamlit as st
import cv2
import gc
from pathlib import Path
from ultralytics import YOLO
import torch
from datetime import datetime
import openpyxl
import pandas as pd
from helper import is_face_present_for_numpy, who, crop_face_from_numpy, extract_location_from_video

# Set Streamlit Page Config
st.set_page_config(page_title="Face Recognition Attendance", layout="wide")

# Set up device (CPU/GPU/MPS)
device = torch.device(
    "mps" if torch.backends.mps.is_available() 
    else "cuda" if torch.cuda.is_available() 
    else "cpu"
)

# Define Paths
attendance_file = Path("outputs/attendance_results/tracked_objects.xlsx")
uploaded_video_dir = Path("data/uploaded_video")
tracking_video_dir = Path("outputs/Tracked_Video")
yolo_model_dir = Path("models/yolo/77_epochs.pt")

# Create directories if not exist
uploaded_video_dir.mkdir(parents=True, exist_ok=True)
tracking_video_dir.mkdir(parents=True, exist_ok=True)

# Load YOLO Model
yolo_model = YOLO(yolo_model_dir)

# Initialize Attendance File
if attendance_file.exists():
    workbook = openpyxl.load_workbook(attendance_file)
    sheet = workbook.active
else:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Attendance Log"
    sheet.append(["Person Name", "Date", "Time", "Last Time", "Location"])
    workbook.save(attendance_file)

# Custom CSS for navigation styling
st.markdown("""
    <style>
    .stRadio > label {
        font-size: 1.2rem;
        font-weight: 500;
        color: #2c3e50;
    }
    .stRadio > div[role='radiogroup'] > label {
        padding: 10px 15px;
        border-radius: 5px;
        margin: 5px 0;
    }
    </style>
""", unsafe_allow_html=True)

# Navigation title in sidebar
st.sidebar.markdown("### 📌 Navigation Menu")

# Navigation radio buttons
page = st.sidebar.radio(
    "",  # Empty label since we have the header above
    ["🏠 Home", "📹 Upload & Play Video", "📜 Attendance"],
    key="nav"
)

# Home Page
if page == "🏠 Home":
    st.markdown("<h1 style='text-align: center; color: #4A90E2;'>🏆 Welcome to the Face Recognition System</h1>", unsafe_allow_html=True)
    st.markdown("<h3 style='text-align: center; color: #333;'>🔍 This system uses <strong>YOLO & AI</strong> to track individuals from videos and log attendance.</h3>", unsafe_allow_html=True)
    
# Upload & Play Video Page
elif page == "📹 Upload & Play Video":
    st.markdown("<h2 style='color: #E67E22;'>🎬 Upload and Track Video</h2>", unsafe_allow_html=True)

    # File Uploader
    uploaded_file = st.file_uploader("📤 Upload a video", type=["mp4", "avi", "mov"])
    
    if uploaded_file is not None:
        # Save uploaded video
        uploaded_video_path = uploaded_video_dir / uploaded_file.name
        with open(uploaded_video_path, "wb") as f:
            f.write(uploaded_file.getbuffer())

        # Display uploaded video
        st.video(str(uploaded_video_path), start_time=0)

        # Extract location
        video_location = extract_location_from_video(uploaded_video_path)
        is_outside = video_location.lower().startswith("of")
        video_exact_location = "🏢 Office Entrance" if is_outside else "📚 Trainee Room"
        st.write(f"📍 Processing video from {video_exact_location}")

        # Start tracking
        with st.spinner("⏳ Processing video..."):
            logged_names = set()
            workbook = openpyxl.load_workbook(attendance_file)
            sheet = workbook.active
            logged_names = {row[0] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]}

            # Video Setup
            cap = cv2.VideoCapture(str(uploaded_video_path))
            fps = int(cap.get(cv2.CAP_PROP_FPS))
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

            # Video Writer Setup
            tracked_video_path = tracking_video_dir / "tracked_video.mp4"
            fourcc = cv2.VideoWriter_fourcc(*"avc1")
            out = cv2.VideoWriter(str(tracked_video_path), fourcc, fps, (width, height))

            frame_count = 0

            while cap.isOpened():
                ret, frame = cap.read()
                if not ret:
                    break

                original_frame = frame.copy()
                results = yolo_model.track(frame, conf=0.5, iou=0.1, persist=True, device=device, imgsz=1024)

                for result in results:
                    for box, obj_id, cls in zip(result.boxes.xyxy, result.boxes.id, result.boxes.cls):
                        x1, y1, x2, y2 = map(int, box)
                        cropped_image = original_frame[y1:y2, x1:x2]

                        if is_face_present_for_numpy(cropped_image, is_outside):
                            cropped_face = crop_face_from_numpy(cropped_image)
                            detected_person = who(cropped_face, is_outside)

                            # Save Cropped Face
                            face_dir = Path("outputs/Tracked_Faces") / detected_person
                            face_dir.mkdir(parents=True, exist_ok=True)
                            cv2.imwrite(str(face_dir / f"ID{int(obj_id)}_frame{frame_count}.jpg"), cropped_face)

                            # Excel Logging
                            current_time = datetime.now()
                            date_str = current_time.strftime("%Y-%m-%d")
                            time_str = current_time.strftime("%H:%M:%S")

                            if detected_person not in logged_names:
                                sheet.append([detected_person, date_str, time_str, time_str, video_exact_location])
                                logged_names.add(detected_person)
                            else:
                                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
                                    if row[0].value == detected_person:
                                        row[3].value = time_str
                                        break

                            # Draw Bounding Box
                            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.putText(
                                frame,
                                f"{detected_person}",
                                (x1, y1 - 10),
                                cv2.FONT_HERSHEY_SIMPLEX,
                                0.5,
                                (0, 255, 0),
                                2,
                            )

                out.write(frame)
                frame_count += 1
                workbook.save(attendance_file)

            # Release Resources
            cap.release()
            out.release()
            cv2.destroyAllWindows()
            gc.collect()

            # Display Processed Video
            if tracked_video_path.exists():
                st.success("✅ Video processing complete!")
                st.video(str(tracked_video_path), start_time=0)
            else:
                st.error("❌ Error: Tracked video was not created!")

# Attendance Page
elif page == "📜 Attendance":
    st.markdown("<h2 style='color: #27AE60;'>📑 Attendance Records</h2>", unsafe_allow_html=True)
    
    if attendance_file.exists():
        df = pd.read_excel(attendance_file)
        st.write(df)
    else:
        st.write("⚠️ No attendance data available yet.")