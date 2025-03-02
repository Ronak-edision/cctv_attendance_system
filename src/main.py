from ultralytics import YOLO
from pathlib import Path
import cv2
import torch
from datetime import datetime
import openpyxl
from helper import is_face_present_for_numpy, who, crop_face_from_numpy, extract_location_from_video

# Device configuration
device = torch.device("mps" if torch.backends.mps.is_available() else 
                      "cuda" if torch.cuda.is_available() else "cpu")

# Initialize Excel logging
excel_path = Path("outputs/attendance_results/tracked_objects.xlsx")
logged_names = set()

# Initialize or load Excel workbook
if excel_path.exists():
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    logged_names = {row[0] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]}
else:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Attendance Log"
    sheet.append(["Person Name", "Date", "Time", "Last Time", "Location"])

# Load YOLO model
yolo_model = YOLO("models/yolo/77_epochs.pt")

# Video setup
video_path = Path("data/uploaded video/Edited_Inferrence_Video.mp4")
cap = cv2.VideoCapture(str(video_path))
video_location = extract_location_from_video(video_path)
print(f"Processing video from {video_location}")
is_outside = not video_location.lower().startswith("tr")
print(f"Is outside: {is_outside}")

# Video writer setup
output_dir = Path("outputs")
tracked_video_path = output_dir / "Tracked Video" / "tracked_video.mp4"
tracked_video_path.parent.mkdir(parents=True, exist_ok=True)
fourcc = cv2.VideoWriter_fourcc(*"avc1")
out = cv2.VideoWriter(str(tracked_video_path), fourcc, 
                     int(cap.get(cv2.CAP_PROP_FPS)), 
                     (int(cap.get(cv2.CAP_PROP_FRAME_WIDTH)), 
                      int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))))

frame_count = 0

while cap.isOpened():
    ret, frame = cap.read()
    if not ret:
        break

    if frame_count % 1 == 0:  # Process every frame
        original_frame = frame.copy()
        results = yolo_model.track(frame, conf=0.5, iou=0.1, 
                                  persist=True, device=device, imgsz=1024)

        for result in results:
            for box, obj_id, cls in zip(result.boxes.xyxy, 
                                      result.boxes.id, 
                                      result.boxes.cls):
                x1, y1, x2, y2 = map(int, box)
                cropped_image = original_frame[y1:y2, x1:x2]

                if is_face_present_for_numpy(cropped_image):
                    cropped_face = crop_face_from_numpy(cropped_image)
                    detected_person = who(cropped_face, is_outside)

                    # Save cropped face
                    face_dir = output_dir / 'Tracked Faces' / detected_person
                    face_dir.mkdir(parents=True, exist_ok=True)
                    cv2.imwrite(str(face_dir / f"ID{int(obj_id)}_frame{frame_count}.jpg"), cropped_face)

                    # Excel logging
                    current_time = datetime.now()
                    date_str = current_time.strftime("%Y-%m-%d")
                    time_str = current_time.strftime("%H:%M:%S")

                    if detected_person not in logged_names:
                        sheet.append([detected_person, date_str, time_str, time_str, video_location])
                        logged_names.add(detected_person)
                    else:
                        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
                            if row[0].value == detected_person:
                                row[3].value = time_str
                                break

                    # Draw bounding box
                    cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                    cv2.putText(frame, f"{detected_person}", (x1, y1 - 10),
                               cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)

        out.write(frame)
        cv2.imshow("YOLO Tracking", frame)

    frame_count += 1
    workbook.save(excel_path)
    if cv2.waitKey(1) & 0xFF == ord("q"):
        break

# Cleanup
cap.release()
out.release()
cv2.destroyAllWindows()